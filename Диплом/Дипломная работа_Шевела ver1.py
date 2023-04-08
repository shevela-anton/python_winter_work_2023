import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import matplotlib
import matplotlib.pyplot as plt
#from matplotlib.backends.backend_qt5 import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import sys
from PyQt6.QtWidgets import (
    QMainWindow, QApplication,
    QLabel, QCheckBox, QComboBox, QLineEdit, QPushButton,
    QLineEdit, QSpinBox, QDoubleSpinBox, QSlider, QHBoxLayout, QGridLayout, QVBoxLayout, QWidget, QSizePolicy
)
from PyQt6.QtGui import QPainter, QPixmap
from PyQt6.QtGui import QStandardItemModel, QStandardItem
from PyQt6.QtCore import Qt
import pandas as pd
import numpy as np
import pyqtgraph as pg
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# Создание основного класса программы
class PersonalAccount(QMainWindow):

    def __init__(self):
        super().__init__()
        #self.tf = True
        #self.text = 'Нажмите Enter!'
        self.setWindowTitle("Личный кабинет")
        self.resize(500, 500)
        self.move(10, 10)

        layout = QGridLayout()

        # Load the Excel file
        workbook = openpyxl.load_workbook('Работа ЛОМ.xlsx')
        worksheet = workbook.active

        # Calculate the total sales amount
        total_sales = 0
        for i in range(1, worksheet.max_row):  # def summ sales
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.':
                total_sales += worksheet.cell(row=i, column=3).value

        # Calculate the total margin
        total_margin = 0
        for i in range(1, worksheet.max_row):  # def summ margin
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.':
                total_margin += worksheet.cell(row=i, column=6).value

        # Create the QComboBox and add it to a QHBoxLayout
        self.month_choose = QComboBox()
        for month in range(1, 13):
            self.month_choose.addItem(f"{month}")
        """self.month_choose.addItems(["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                               "Июль", "Август", "Сентябрь", "Ноябрь", "Декабрь"])"""
        self.month_choose.currentIndexChanged.connect(self.calculate_salary)
        layout.addWidget(self.month_choose, 8, 5)

        # Создаём QLabel для вывода суммы премии
        self.salary_label = QLabel(self)
        font = self.salary_label.font()
        font.setPointSize(10)
        self.salary_label.setFont(font)
        layout.addWidget(self.salary_label, 10, 7)

        # Создаём QLabel для вывода информации о сотруднике
        position = QLabel("Должность")
        font = position.font()
        font.setPointSize(10)
        position.setFont(font)
        layout.addWidget(position, 2, 2)

        name = QLabel("ФИО")
        font = name.font()
        font.setPointSize(10)
        name.setFont(font)
        layout.addWidget(name, 3, 2)

        contacts = QLabel("Контактная информация")
        font = contacts.font()
        font.setPointSize(10)
        contacts.setFont(font)
        layout.addWidget(contacts, 1, 5)

        manager = QLabel("Менеджер по продажам")
        font = manager.font()
        font.setPointSize(10)
        manager.setFont(font)
        layout.addWidget(manager, 2, 3)

        name_label = QLabel("Кривоносова Елена Андреевна")
        font = name_label.font()
        font.setPointSize(10)
        name_label.setFont(font)
        layout.addWidget(name_label, 3, 3)

        email = QLabel("email")
        font = email.font()
        font.setPointSize(10)
        email.setFont(font)
        layout.addWidget(email, 2, 5)

        phone = QLabel("Рабочий телефон")
        font = phone.font()
        font.setPointSize(10)
        phone.setFont(font)
        layout.addWidget(phone, 3, 5)

        email_company = QLabel("i.ivanov@company.ru")
        font = email_company.font()
        font.setPointSize(10)
        email_company.setFont(font)
        layout.addWidget(email_company, 2, 6)

        mob_phone = QLabel("+ 7 911 700-00-07")
        font = mob_phone.font()
        font.setPointSize(10)
        mob_phone.setFont(font)
        layout.addWidget(mob_phone, 3, 6)

        key_customers = QLabel("Ключевые клиенты")
        font = key_customers.font()
        font.setPointSize(10)
        key_customers.setFont(font)
        layout.addWidget(key_customers, 1, 8)

        customers = QLabel("!!!!!!!!!!Ключевые клиенты!!!!!")
        font = customers.font()
        font.setPointSize(10)
        customers.setFont(font)
        layout.addWidget(customers, 2, 8, 2, 2)

        plan_comp = QLabel("Выполнено")
        font = plan_comp.font()
        font.setPointSize(10)
        plan_comp.setFont(font)
        layout.addWidget(plan_comp, 5, 11)

        percent = QLabel("%")
        font = percent.font()
        font.setPointSize(10)
        percent.setFont(font)
        layout.addWidget(percent, 5, 12)

        year_plan = QLabel("Годовой план")
        font = year_plan.font()
        font.setPointSize(10)
        year_plan.setFont(font)
        layout.addWidget(year_plan, 6, 9)

        quater_plan = QLabel("Квартальный план")
        font = quater_plan.font()
        font.setPointSize(10)
        quater_plan.setFont(font)
        layout.addWidget(quater_plan, 7, 9)

        plan_amount = QLabel("100 000 000")
        font = plan_amount.font()
        font.setPointSize(10)
        plan_amount.setFont(font)
        layout.addWidget(plan_amount, 6, 10)

        quater_plan_amount = QLabel("30 000 000")
        font = quater_plan_amount.font()
        font.setPointSize(10)
        quater_plan_amount.setFont(font)
        layout.addWidget(quater_plan_amount, 7, 10)

        year_plan_res = QLabel("{:10.2f}".format(total_sales))
        font = year_plan_res.font()
        font.setPointSize(10)
        year_plan_res.setFont(font)
        layout.addWidget(year_plan_res, 6, 11)

        quater_plan_res = QLabel(f"{49166570.29}")
        font = quater_plan_res.font()
        font.setPointSize(10)
        quater_plan_res.setFont(font)
        layout.addWidget(quater_plan_res, 7, 11)

        percent_res = QLabel("{:<10.0f}".format(total_sales/100000000*100))
        font = percent_res.font()
        font.setPointSize(10)
        percent_res.setFont(font)
        layout.addWidget(percent_res, 6, 12)

        percent_quater_res = QLabel("{:<10.0f}".format(49166570.29/30000000*100))
        font = percent_quater_res.font()
        font.setPointSize(10)
        percent_quater_res.setFont(font)
        layout.addWidget(percent_quater_res, 7, 12)




        # Построение графика
        self.graphic = QLabel(self)
        layout.addWidget(self.graphic, 13, 8)
        # Получение значений для графика
        lst_month = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        lst_sales = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for i in range(2, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.':
                lst_sales[worksheet.cell(row=i, column=4).value.month - 1] += worksheet.cell(row=i, column=3).value


        fig, ax = plt.subplots()
        ax.bar(lst_month, lst_sales)
        ax.set_xlabel("Месяц")
        ax.set_ylabel("Сумма сделок")
        ax.set_title("Размещённые заказы за год")

        # Save the plot to a PNG file
        fig.savefig("sales_by_month.png")

        # Display the plot in the GUI
        pixmap = QPixmap("sales_by_month.png")
        self.graphic.setPixmap(pixmap)


        statistics = QLabel("Статистика сотрудника")
        font = statistics.font()
        font.setPointSize(10)
        statistics.setFont(font)
        layout.addWidget(statistics, 8, 1)

        calls = QLabel("Количество звонков")
        font = calls.font()
        font.setPointSize(10)
        calls.setFont(font)
        layout.addWidget(calls, 9, 1)

        mails = QLabel("Количество писем")
        font = mails.font()
        font.setPointSize(10)
        mails.setFont(font)
        layout.addWidget(mails, 10, 1)

        meetings = QLabel("Количество встреч")
        font = meetings.font()
        font.setPointSize(10)
        meetings.setFont(font)
        layout.addWidget(meetings, 11, 1)

        trips = QLabel("Планируемые командировки")
        font = trips.font()
        font.setPointSize(10)
        trips.setFont(font)
        layout.addWidget(trips, 12, 1)

        trips_list = QLabel("!!!!командировки!!!!")
        font = trips_list.font()
        font.setPointSize(10)
        trips_list.setFont(font)
        layout.addWidget(trips_list, 13, 1, 2, 2)

        month = QLabel("Выбор месяца")
        font = month.font()
        font.setPointSize(10)
        month.setFont(font)
        layout.addWidget(month, 8, 4)


        new_contracts = QLabel("Заключенные сделки")
        font = new_contracts.font()
        font.setPointSize(10)
        new_contracts.setFont(font)
        layout.addWidget(new_contracts, 9, 5)

        closed_contracts = QLabel("Закрытые сделки")
        font = closed_contracts.font()
        font.setPointSize(10)
        closed_contracts.setFont(font)
        layout.addWidget(closed_contracts, 10, 5)

        self.new_summ = QLabel(self)
        font = self.new_summ.font()
        font.setPointSize(10)
        self.new_summ.setFont(font)
        layout.addWidget(self.new_summ, 9, 6)

        self.closed_summ = QLabel(self)
        font = self.closed_summ.font()
        font.setPointSize(10)
        self.closed_summ.setFont(font)
        layout.addWidget(self.closed_summ, 10, 6)

        salary_summ = QLabel("Сумма премии")
        font = salary_summ.font()
        font.setPointSize(10)
        salary_summ.setFont(font)
        layout.addWidget(salary_summ, 9, 7)



        self.salary_button = QPushButton("Рассчитать ЗП")
        #salary_button.setCheckable(True)
        #button.clicked.connect(self.the_button_was_clicked)
        self.salary_button.clicked.connect(self.calculate_salary)
        layout.addWidget(self.salary_button, 8, 6)

        businesstrip_button = QPushButton("Сформировать запрос на командировку")
        #businesstrip_button.setCheckable(True)
        layout.addWidget(businesstrip_button, 12, 3)

        absence_button = QPushButton("Сформировать согласование на отсутствие")
       # absence_button.setCheckable(True)
        layout.addWidget(absence_button, 13, 3)


        widgets = [position, name, contacts, manager, name_label, email, phone, email_company, mob_phone, key_customers,
                   customers, plan_comp, percent, year_plan, quater_plan, plan_amount, quater_plan_amount,
                   year_plan_res, quater_plan_res, percent_res, percent_quater_res, self.graphic, statistics,
                   calls, mails, meetings, trips, trips_list, month, self.month_choose, new_contracts, closed_contracts,
                   self.new_summ,
                   self.closed_summ, salary_summ, self.salary_button, businesstrip_button, absence_button]
        for w in widgets:
            layout.addWidget(w)

        widgets = QWidget()
        widgets.setLayout(layout)
        self.setCentralWidget(widgets)



    def calculate_salary(self):
        # Load the Excel file
        workbook = openpyxl.load_workbook('Работа ЛОМ.xlsx')
        worksheet = workbook.active

        # Get the selected month
        month = int(self.month_choose.currentText())

        # Calculate the total sales amount for the selected month
        lst_sales = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        total_sales = 0
        for i in range(1, worksheet.max_row):  # def summ sales
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.' \
                    and worksheet.cell(row=i, column=4).value.month == month:
                total_sales += worksheet.cell(row=i, column=3).value
                lst_sales[worksheet.cell(row=i, column=4).value.month - 1] += worksheet.cell(row=i, column=3).value
        self.new_summ.setText(f"{total_sales}")



        # Calculate the total margin for the selected month
        total_margin = 0
        for i in range(2, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.' and \
                    worksheet.cell(row=i, column=4).value.month == month \
                    and worksheet.cell(row=i, column=7).value.year == 2022:
                total_margin += worksheet.cell(row=i, column = 6).value
        self.closed_summ.setText(f"{total_margin}")


        # Calculate the salary
        salary = (total_sales * 0.0017) + (total_margin * 0.03)

        # Update the QLabels with the new salary information
        self.salary_label.setText("{:<10.0f}".format(salary))



if __name__ == '__main__':
    # Create the PyQt6 application and PersonalAccount widget
    app = QApplication(sys.argv)
    w = PersonalAccount()
    w.show()
    app.exec()

