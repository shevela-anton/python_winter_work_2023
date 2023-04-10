import openpyxl
import matplotlib.pyplot as plt
import sys
from PyQt6.QtWidgets import (
    QMainWindow, QApplication, QScrollArea,
    QLabel, QComboBox, QPushButton, QVBoxLayout,
    QHBoxLayout, QGridLayout, QWidget
)
from PyQt6.QtGui import QPixmap
from PyQt6.QtCore import Qt


# Создание основного класса программы
class PersonalAccount(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Личный кабинет")
        self.resize(1280, 720)
        self.move(0, 0)

        # Для разных размеров экрана устанавливаем прокрутку
        # Так как дизайн верстался на мониторе 29 дюймов в 4к
        self.scroll = QScrollArea()
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.scroll.setWidgetResizable(True)

        # Создаём главный layout
        layout = QGridLayout()
        layout.setContentsMargins(20, 20, 20, 140)

        # Создаём вложенные слои
        layout_top = QGridLayout()
        layout_left = QVBoxLayout()
        layout_personal = QGridLayout()
        layout_sal = QGridLayout()
        layout_salper = QGridLayout()
        layout_head = QHBoxLayout()
        layout_customers = QGridLayout()
        layout_progress = QGridLayout()
        layout_trips = QGridLayout()
        layout_graph = QGridLayout()

        # Устанавливаем границы слоёв
        layout_top.setSpacing(0)
        layout_head.setSpacing(28)
        layout_personal.setContentsMargins(40, 40, 36, 40)
        layout_sal.setContentsMargins(40, 40, 100, 40)
        layout_sal.setContentsMargins(40, 40, 40, 40)
        layout_customers.setContentsMargins(40, 40, 40, 40)

        # Вкладываем слои в главный layout
        layout_salper.addLayout(layout_personal, 1, 0)
        layout_salper.addLayout(layout_sal, 2, 0)
        layout.addLayout(layout_salper, 3, 1)
        layout.addLayout(layout_left, 1, 0, 5, 1)
        layout.addLayout(layout_top, 0, 0, 1, 4)
        layout.addLayout(layout_head, 1, 0, 1, 4)
        layout.addLayout(layout_customers, 3, 2)
        layout.addLayout(layout_progress, 3, 3)
        layout.addLayout(layout_graph, 4, 3)
        layout.addLayout(layout_trips, 4, 2)

        # Загрузка Excel файла
        workbook = openpyxl.load_workbook('Работа ЛОМ.xlsx')
        worksheet = workbook.active

        # Рассчитывает общую сумму продаж по выбранному менеджеру
        total_sales = 0
        for i in range(1, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.':
                total_sales += worksheet.cell(row=i, column=3).value

        # Рассчитывает общую сумму прибыли по выбранному менеджеру
        total_margin = 0
        for i in range(1, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.':
                total_margin += worksheet.cell(row=i, column=6).value

        # Создаём выбор месяца QComboBox и добавляем к интерфейсу QHBoxLayout
        self.month_choose = QComboBox()
        for month in range(1, 13):
            self.month_choose.addItem(f"{month}")
        self.month_choose.setStyleSheet("QLabel {width: 108px; height: 20px; font-family: 'SF Pro Display'; "
                                        "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px;"
                                        "color: #2A4FFF; flex: none; order: 1; flex-grow: 0;}")
        # Добавляем связь с функцией расчёта зарплаты
        self.month_choose.currentIndexChanged.connect(self.calculate_salary)
        layout_sal.addWidget(self.month_choose, 1, 2)

        # Создаём Tab
        tab = QLabel()
        pixmap = QPixmap("Tab.png")
        tab.setPixmap(pixmap)
        tab.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        tab.setStyleSheet("QLabel {background-color: #F7F7FA}")
        layout_left.addWidget(tab)

        # Создаём шапку QLabel
        top_label = QLabel()
        pixmap = QPixmap("Head.png")
        top_label.setPixmap(pixmap)
        top_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        top_label.setStyleSheet("QLabel {background-color: #F7F7FA}")
        layout_top.addWidget(top_label, 1, 1)

        top_label1 = QLabel()
        pixmap = QPixmap("Head_r.png")
        top_label1.setPixmap(pixmap)
        top_label1.setAlignment(Qt.AlignmentFlag.AlignRight)
        top_label1.setStyleSheet("QLabel {background-color: #F7F7FA}")
        layout_top.addWidget(top_label1, 1, 4)

        # Создаём QLabel Рассчитать зарплату
        salary_panel = QLabel('Рассчитать зарплату')
        salary_panel.setStyleSheet("QLabel {width: 214px; height: 24px; font-family: 'SF Pro Display'; "
                                   "font-style: normal; font-weight: 700; font-size: 21px; line-height: 24px;"
                                   "color: #29292C; flex: none; order: 0; flex-grow: 0;}")
        layout_sal.addWidget(salary_panel, 0, 1)

        # Создаём QLabel для вывода суммы премии
        self.salary_label = QLabel(self)
        self.salary_label.setStyleSheet("QLabel {width: 14px; height: 24px; font-family: 'SF Pro Display';"
                                        "font-style: normal; font-weight: 700; font-size: 21px;line-height: 24px; "
                                        "color: #29292C; flex: none; order: 0; flex-grow: 0;}")

        layout_sal.addWidget(self.salary_label, 2, 1)

        # Создаём QLabel для вывода информации о сотруднике
        contacts = QLabel("Личная информация")
        contacts.setStyleSheet("QLabel {width: 170px; height: 24px; font-family: 'SF Pro Display'; "
                               "font-style: normal; font-weight: 500; font-size: 18px; line-height: 24px; "
                               "color: #29292C; flex: none; order: 0; flex-grow: 0;}")
        layout_personal.addWidget(contacts, 1, 1)

        position = QLabel("Должность")
        position.setStyleSheet("QLabel {width: 92px; height: 20px; font-family: 'SF Pro Display'; "
                               "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px;"
                               "color: #2A4FFF; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(position, 2, 1)

        name = QLabel("ФИО")
        name.setStyleSheet("QLabel {width: 92px; height: 20px; font-family: 'SF Pro Display'; "
                           "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px;"
                           "color: #2A4FFF; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(name, 3, 1)

        manager = QLabel("Менеджер по продажам")
        manager.setStyleSheet("QLabel {width: 252px; height: 20px; font-family: 'SF Pro Display'; "
                              "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px; "
                              "color: #31435A; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(manager, 2, 2)

        name_label = QLabel("Кривоносова Елена Андреевна")
        name_label.setStyleSheet("QLabel {width: 252px; height: 20px; font-family: 'SF Pro Display'; "
                                 "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px; "
                                 "color: #31435A; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(name_label, 3, 2)

        email = QLabel("Email")
        email.setStyleSheet("QLabel {width: 92px; height: 20px; font-family: 'SF Pro Display'; "
                            "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px;"
                            "color: #2A4FFF; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(email, 4, 1)

        phone = QLabel("Телефон")
        phone.setStyleSheet("QLabel {width: 92px; height: 20px; font-family: 'SF Pro Display'; "
                            "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px;"
                            "color: #2A4FFF; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(phone, 5, 1)

        email_company = QLabel("e.krivonosova@company.ru")
        email_company.setStyleSheet("QLabel {width: 252px; height: 20px; font-family: 'SF Pro Display'; "
                                    "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px; "
                                    "color: #31435A; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(email_company, 4, 2)

        mob_phone = QLabel("+ 7 911 700-00-07")
        mob_phone.setStyleSheet("QLabel {width: 252px; height: 20px; font-family: 'SF Pro Display'; "
                                "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px; "
                                "color: #31435A; flex: none; order: 1; flex-grow: 0;}")
        layout_personal.addWidget(mob_phone, 5, 2)

        # Создаём блок ключевых клиентов
        key_customers = QLabel("Ключевые клиенты")
        key_customers.setStyleSheet("QLabel {width: 158px;height: 24px; font-family: 'SF Pro Display'; "
                                    "font-weight: 500; font-size: 18px; line-height: 24px; color: #29292C;"
                                    "flex: none; order: 0; flex-grow: 0;}")
        layout_customers.addWidget(key_customers, 1, 1)

        # Здесь будет вывод ключевых клиентов из загружаемого файла
        customers = QLabel()
        pixmap = QPixmap("Клиенты.png")
        customers.setPixmap(pixmap)
        customers.setStyleSheet("QLabel {box-sizing: border-box; width: 556px; height: 208px; "
                                "border-bottom: 1px solid #CBD1DF; flex: none; order: 1; flex-grow: 0;}")
        customers.setAlignment(Qt.AlignmentFlag.AlignLeft)
        customers.resize(556, 208)
        layout_customers.addWidget(customers, 2, 1, 3, 2)

        # Создаём блок выполнения плана
        plan_comp = QLabel("Прогресс")
        plan_comp.setStyleSheet("QLabel {width: 80px; height: 24px; font-family: 'SF Pro Display'; "
                                "font-style: normal; font-weight: 500; font-size: 18px; line-height: 24px;"
                                "color: #29292C; flex: none; order: 0; flex-grow: 0;}")
        layout_progress.addWidget(plan_comp, 1, 1)

        # Блок процента выполнения плана

        year_plan = QLabel("Годовой план")
        year_plan.setStyleSheet("QLabel {width: 87px; height: 16px; font-family: 'SF Pro Display'; "
                                "font-style: normal; font-weight: 500; font-size: 14px; line-height: 16px;"
                                "color: #31435A; flex: none; order: 0; flex-grow: 0;}")
        layout_progress.addWidget(year_plan, 3, 1)

        quater_plan = QLabel("Квартальный план")
        quater_plan.setStyleSheet("QLabel {width: 87px; height: 16px; font-family: 'SF Pro Display'; "
                                  "font-style: normal; font-weight: 500; font-size: 14px; line-height: 16px;"
                                  "color: #31435A; flex: none; order: 0; flex-grow: 0;}")
        layout_progress.addWidget(quater_plan, 3, 2)

        plan_amount = QLabel("100 млн")
        plan_amount.setStyleSheet("QLabel {width: 53px; height: 16px; font-family: 'SF Pro Display';"
                                  "font-style: normal; font-weight: 500; font-size: 14px; line-height: 16px; "
                                  "color: #8B8D97; flex: none; order: 1; flex-grow: 0;}")
        layout_progress.addWidget(plan_amount, 4, 1)

        quater_plan_amount = QLabel("30 млн")
        quater_plan_amount.setStyleSheet("QLabel {width: 53px; height: 16px; font-family: 'SF Pro Display'; "
                                         "font-style: normal; font-weight: 500; font-size: 14px; line-height: 16px; "
                                         "color: #8B8D97; flex: none; order: 1; flex-grow: 0;}")
        layout_progress.addWidget(quater_plan_amount, 4, 2)

        # Создаём вывод выполненной суммы плана
        year_plan_res = QLabel("Выполнено: " + "{:10.0f}".format(total_sales//1000000) + " млн")
        year_plan_res.setStyleSheet("QLabel {position: absolute; width: 149px; height: 20px; left: 47px; top: 64px;"
                                    "font-family: 'SF Pro Display'; font-style: normal; font-weight: 500; "
                                    "font-size: 16px; line-height: 20px; color: #8B8D97;}")
        layout_progress.addWidget(year_plan_res, 2, 1)

        quater_plan_res = QLabel("Выполнено: 49 млн")
        quater_plan_res.setStyleSheet("QLabel {position: absolute; width: 149px; height: 20px; left: 47px; top: 64px;"
                                      "font-family: 'SF Pro Display'; font-style: normal; font-weight: 500;"
                                      "font-size: 16px; line-height: 20px; color: #8B8D97;}")
        layout_progress.addWidget(quater_plan_res, 2, 2)

        # Вывод диаграммы
        percent_res = QLabel()
        pixmap = QPixmap("Диаграмма 165.png")
        percent_res.setPixmap(pixmap)
        percent_res.setStyleSheet("QLabel {position: absolute; width: 143px; height: 143px; left: 47px; top: 162px;}")
        percent_res.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_progress.addWidget(percent_res, 5, 1)

        percent_quater_res = QLabel()
        pixmap = QPixmap("Диаграмма 164.png")
        percent_quater_res.setPixmap(pixmap)
        percent_quater_res.setStyleSheet("QLabel {position: absolute; width: 143px; height: 143px; left: 263px; "
                                         "top: 162px;}")
        percent_quater_res.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_progress.addWidget(percent_quater_res, 5, 2)

        # Построение графика
        self.graphic = QLabel(self)
        layout_graph.addWidget(self.graphic, 1, 1)
        # Получение значений для графика
        lst_month = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        lst_sales = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for i in range(2, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.':
                lst_sales[worksheet.cell(row=i, column=4).value.month - 1] += \
                    worksheet.cell(row=i, column=3).value/1000000

        fig, ax = plt.subplots()
        ax.bar(lst_month, lst_sales, color='#6587FF')
        ax.set_xlabel("Месяц")
        ax.set_ylabel("Сумма сделок, млн руб")
        ax.set_title("Размещённые заказы за год")

        # Сохраняем график в виде PNG файла
        fig.savefig("sales_by_month.png")

        # Отображение графика в GUI
        pixmap = QPixmap("sales_by_month.png")
        self.graphic.setPixmap(pixmap)

        # Создание блока статистики
        calls = QLabel("Количество звонков: <b> 120 <b>")
        calls.setStyleSheet("QLabel {display: flex; flex-direction: row; justify-content: center; align-items: center; "
                            "padding: 14px 24px; gap: 24px; width: 278px; height: 48px; background: #FFFFFF;"
                            "box-shadow: 0px 4px 13px rgba(187, 196, 227, 0.32); border-radius: 25px; flex: none;"
                            "order: 0; flex-grow: 0; font-family: 'SF Pro Display'; font-style: normal; "
                            "font-weight: 500; font-size: 16px; line-height: 20px; color: #8B8D97;}")
        layout_head.addWidget(calls)

        mails = QLabel("Количество писем: <b> 34 <b>")
        mails.setStyleSheet("QLabel {display: flex; flex-direction: row; justify-content: center; align-items: center; "
                            "padding: 14px 24px; gap: 24px; width: 278px; height: 48px; background: #FFFFFF;"
                            "box-shadow: 0px 4px 13px rgba(187, 196, 227, 0.32); border-radius: 25px; flex: none;"
                            "order: 0; flex-grow: 0; font-family: 'SF Pro Display'; font-style: normal; "
                            "font-weight: 500; font-size: 16px; line-height: 20px; color: #8B8D97;}")
        layout_head.addWidget(mails)

        meetings = QLabel("Количество встреч: <b> 10 <b>")
        meetings.setStyleSheet("QLabel {display: flex; flex-direction: row; justify-content: center; "
                               "align-items: center; padding: 14px 24px; gap: 24px; width: 278px; height: 48px; "
                               "background: #FFFFFF; box-shadow: 0px 4px 13px rgba(187, 196, 227, 0.32); "
                               "border-radius: 25px; flex: none; order: 0; flex-grow: 0; font-family: 'SF Pro Display';"
                               "font-style: normal; font-weight: 500; font-size: 16px; line-height: 20px; "
                               "color: #8B8D97; flex: none; order: 0; flex-grow: 0;}")
        layout_head.addWidget(meetings)

        # Создание блока командировок
        trips = QLabel("Планируемые командировки")
        trips.setStyleSheet("QLabel {width: 241px; height: 24px; font-family: 'SF Pro Display'; font-style: normal; "
                            "font-weight: 500; font-size: 18px; line-height: 24px; color: #29292C; flex: none; "
                            "order: 0; flex-grow: 0;}")
        layout_trips.addWidget(trips, 1, 1)

        trips_list = QLabel()
        pixmap = QPixmap("Командировки.png")
        trips_list.setPixmap(pixmap)
        trips_list.setStyleSheet("display: flex; flex-direction: column; align-items: flex-start; padding: 0px; "
                                 "width: 557px; height: 120px; flex: none; order: 1; align-self: stretch; "
                                 "flex-grow: 0;}")
        trips_list.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        layout_trips.addWidget(trips_list, 2, 1, 1, 2)

        # Создание блока вывода информации о зарплате
        month = QLabel("Выбор месяца")
        month.setStyleSheet("QLabel {width: 108px; height: 20px; font-family: 'SF Pro Display'; "
                            "font-style: normal; font-weight: 400; font-size: 16px; line-height: 20px;"
                            "color: #2A4FFF; flex: none; order: 1; flex-grow: 0;}")
        layout_sal.addWidget(month, 1, 1)

        new_contracts = QLabel("Заключенные сделки")
        new_contracts.setStyleSheet("QLabel {width: 129px; height: 20px; font-family: 'SF Pro Display'; "
                                    "font-style: normal; font-weight: 500; font-size: 16px; line-height: 20px;"
                                    "color: #7A7D8B; flex: none; order: 1; flex-grow: 0;}")
        layout_sal.addWidget(new_contracts, 6, 1)

        closed_contracts = QLabel("Закрытые сделки")
        closed_contracts.setStyleSheet("QLabel {width: 129px; height: 20px; font-family: 'SF Pro Display'; "
                                       "font-style: normal; font-weight: 500; font-size: 16px; line-height: 20px;"
                                       "color: #7A7D8B; flex: none; order: 1; flex-grow: 0;}")
        layout_sal.addWidget(closed_contracts, 6, 2)

        self.new_summ = QLabel(self)
        self.new_summ.setStyleSheet("QLabel {width: 13px; height: 28px; font-family: 'SF Pro Display'; "
                                    "font-style: normal; font-weight: 500; font-size: 20px; line-height: 28px; "
                                    "color: #29292C; flex: none; order: 0; flex-grow: 0;}")
        layout_sal.addWidget(self.new_summ, 5, 1)

        self.closed_summ = QLabel(self)
        self.closed_summ.setStyleSheet("QLabel {width: 13px; height: 28px; font-family: 'SF Pro Display'; "
                                       "font-style: normal; font-weight: 500; font-size: 20px; line-height: 28px; "
                                       "color: #29292C; flex: none; order: 0; flex-grow: 0;}")
        layout_sal.addWidget(self.closed_summ, 5, 2)

        salary_summ = QLabel("Сумма премии")
        salary_summ.setStyleSheet("QLabel {width: 110px; height: 20px; font-family: 'SF Pro Display'; "
                                  "font-style: normal; font-weight: 500; font-size: 16px; line-height: 20px; "
                                  "color: #7A7D8B; flex: none; order: 1; flex-grow: 0;}")
        layout_sal.addWidget(salary_summ, 3, 1)

        # Создание кнопки расчета зарплаты и связи с функцией расчёта
        self.salary_button = QPushButton("Рассчитать ЗП")
        self.salary_button.setCheckable(True)
        self.salary_button.clicked.connect(self.calculate_salary)
        self.salary_button.setStyleSheet("QPushButton {display: flex; flex-direction: row; justify-content: center; "
                                         "align-items: center; padding: 8px 12px 8px 16px; gap: 4px; width: 195px; "
                                         "height: 40px; background: #2A4FFF; border-radius: 28px; }")
        layout_sal.addWidget(self.salary_button, 2, 2)

        # Создание кнопки запроса на согласование командировки
        businesstrip_button = QPushButton("Сформировать запрос")
        layout_trips.addWidget(businesstrip_button, 1, 2)
        businesstrip_button.setStyleSheet("QPushButton {display: flex; flex-direction: row; justify-content: center; "
                                          "align-items: center; padding: 10px 16px 10px 20px; gap: 4px; "
                                          "position: absolute; width: 262px; height: 40px; right: 36px; top: 160px; "
                                          "background: #6771A2; border-radius: 28px;}")

        # Создание кнопки запроса на согласование отсутствия
        absence_button = QPushButton("Cогласование отсутствия")
        layout_head.addWidget(absence_button, 1, alignment=Qt.AlignmentFlag.AlignRight)
        absence_button.setStyleSheet("QPushButton {display: flex; flex-direction: row; justify-content: center; "
                                     "align-items: center; padding: 10px 16px 10px 20px; gap: 4px; position: absolute;"
                                     "width: 262px; height: 40px; right: 36px; top: 160px; background: #6771A2; "
                                     "border-radius: 28px;}")

        widgets = QWidget()
        widgets.setLayout(layout)
        self.scroll.setWidget(widgets)

        self.setCentralWidget(self.scroll)

    def calculate_salary(self):
        # Загрузить Excel файл
        workbook = openpyxl.load_workbook('Работа ЛОМ.xlsx')
        worksheet = workbook.active

        # Получаем выбранный месяц
        month = int(self.month_choose.currentText())

        # Расчёт общей суммы продаж за выбранный месяц
        lst_sales = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        total_sales = 0
        for i in range(1, worksheet.max_row):  # def summ sales
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.' \
                    and worksheet.cell(row=i, column=4).value.month == month:
                total_sales += worksheet.cell(row=i, column=3).value
                lst_sales[worksheet.cell(row=i, column=4).value.month - 1] += worksheet.cell(row=i, column=3).value
        self.new_summ.setText(f"{total_sales}")

        # Расчёт общей суммы прибыли за выбранный месяц
        total_margin = 0
        for i in range(2, worksheet.max_row):
            if worksheet.cell(row=i, column=1).value == 'Кривоносова Е.' and \
                    worksheet.cell(row=i, column=4).value.month == month \
                    and worksheet.cell(row=i, column=7).value.year == 2022:
                total_margin += worksheet.cell(row=i, column=6).value
        self.closed_summ.setText(f"{total_margin}")

        # Расчёт зарплаты
        salary = (total_sales * 0.0017) + (total_margin * 0.03)

        # Обновление QLabels новой информацией по зарплате
        self.salary_label.setText("{:<10.0f}".format(salary))


if __name__ == '__main__':
    # Создание PyQt6 приложения и PersonalAccount widget
    app = QApplication(sys.argv)
    w = PersonalAccount()

    app.setStyleSheet("""
        QWidget {
        background-color: #F7F7FA;
        }
        QLabel {
        font-family: 'SF Pro Display';
        background-color: #FFFFFF;
        box-shadow: 0px 4px 24px rgba(160, 167, 192, 0.13);
        border-radius: 20px;
        }
        QPushButton {  
            font-family: 'SF Pro Display';
            font-style: normal;
            font-weight: 500;
            font-size: 16px;
            line-height: 20px;
            color: #FFFFFF; 
        }
        QComboBox {
            font-family: 'SF Pro Display';
            background-color: #FFFFFF;
            color: "#2A4FFF";
        }
    """)

    w.show()
    app.exec()
